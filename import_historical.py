import openpyxl
import requests
import json
import os
import time
import signal
from bs4 import BeautifulSoup
from datetime import datetime, timezone

INPUT_EXCEL  = "OJPH(2).xlsx"
OUTPUT_FILE  = "enriched_jobs.json"
DELAY        = 1
MAX_RETRIES  = 3
SAVE_EVERY   = 20

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Referer": "https://www.onlinejobs.ph/",
}

_interrupted = False

def _on_interrupt(signum, frame):
    global _interrupted
    _interrupted = True
    print("\n  Interrupt received — saving and exiting...")

signal.signal(signal.SIGINT, _on_interrupt)
signal.signal(signal.SIGTERM, _on_interrupt)


def load_json(filepath):
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []
    return []


def save_json(filepath, data):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def extract_job_id(url):
    try:
        return url.rstrip("/").split("/")[-1]
    except Exception:
        return None


def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    urls_seen = set()
    rows = []

    for r in range(2, ws.max_row + 1):
        url = ws.cell(row=r, column=7).value
        if not url:
            continue
        url = str(url).strip()
        if url in urls_seen:
            continue
        urls_seen.add(url)

        rows.append({
            "url": url,
            "contact_person": str(ws.cell(row=r, column=1).value or "").strip(),
            "work_type_excel": str(ws.cell(row=r, column=2).value or "").strip(),
            "salary_excel": str(ws.cell(row=r, column=3).value or "").strip(),
            "hours_excel": str(ws.cell(row=r, column=4).value or "").strip(),
            "date_updated_excel": str(ws.cell(row=r, column=5).value or "").strip(),
            "message_excel": str(ws.cell(row=r, column=6).value or "").strip(),
            "notes": str(ws.cell(row=r, column=8).value or "").strip(),
            "source_info": str(ws.cell(row=r, column=9).value or "").strip(),
            "status": str(ws.cell(row=r, column=10).value or "").strip(),
            "contact_info": str(ws.cell(row=r, column=11).value or "").strip(),
        })

    return rows


def fetch_page(url):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = requests.get(url, headers=HEADERS, timeout=30)
            if response.status_code == 200 and "onlinejobs" in response.text.lower():
                return response.text
            elif response.status_code in (404, 410):
                return None
            else:
                if attempt < MAX_RETRIES:
                    time.sleep(2)
        except requests.RequestException:
            if attempt < MAX_RETRIES:
                time.sleep(2)
    return None


def parse_detail_page(html):
    soup = BeautifulSoup(html, "html.parser")

    job_title = "N/A"
    h1 = soup.select_one("h1.fs-24.fw-600.text-white.text-center.mb-40.job__title")
    if h1:
        job_title = h1.get_text(strip=True)

    work_type = "N/A"
    salary_detail = "N/A"
    hours = "N/A"
    date_updated = "N/A"

    card = soup.select_one("div.card.job-post")
    if card:
        values = [p.get_text(strip=True) for p in card.select("p.fs-18")]
        if len(values) >= 4:
            work_type = values[0] or "N/A"
            salary_detail = values[1] or "N/A"
            hours = values[2] or "N/A"
            date_updated = values[3] or "N/A"

    description = "N/A"
    desc_tag = soup.select_one("p#job-description") or soup.select_one("p.job-description")
    if desc_tag:
        description = desc_tag.get_text(separator=" ", strip=True)

    return {
        "job_title": job_title,
        "work_type": work_type,
        "salary_detail": salary_detail,
        "hours": hours,
        "date_updated": date_updated,
        "description": description,
    }


def build_enriched_job(excel_row, detail, scraped_ok):
    job_id = extract_job_id(excel_row["url"])
    now = datetime.now(timezone.utc).isoformat()

    if scraped_ok and detail:
        title    = detail["job_title"]
        job_title_d = detail["job_title"]
        work_type    = detail["work_type"]
        salary_d     = detail["salary_detail"]
        hours        = detail["hours"]
        date_updated = detail["date_updated"]
        description  = detail["description"]
    else:
        title        = "N/A"
        job_title_d = excel_row["message_excel"][:100] if excel_row["message_excel"] else "N/A"
        work_type    = excel_row["work_type_excel"] or "N/A"
        salary_d     = excel_row["salary_excel"] or "N/A"
        hours        = excel_row["hours_excel"] or "N/A"
        date_updated = excel_row["date_updated_excel"] or "N/A"
        description  = excel_row["message_excel"] or "N/A"

    return {
        "job_id":            job_id,
        "title":             title,
        "company":           "N/A",
        "type":              work_type,
        "salary":            salary_d,
        "posted":            excel_row["date_updated_excel"] or "N/A",
        "url":               excel_row["url"],
        "tags":              [],
        "scraped_at":        now,
        "job_title":         job_title_d,
        "work_type":         work_type,
        "salary_detail":     salary_d,
        "hours":             hours,
        "date_updated":      date_updated,
        "description":       description,
        "company_detail":    "N/A",
        "contact_person":    excel_row["contact_person"] or "N/A",
        "detail_scraped_at": now,
        "notes":             excel_row["notes"],
        "source_info":       excel_row["source_info"],
        "status":            excel_row["status"],
        "contact_info":      excel_row["contact_info"],
        "email":             "",
        "extra_description": "",
    }


def main():
    print("Reading Excel...")
    excel_rows = read_excel(INPUT_EXCEL)
    print(f"  {len(excel_rows)} unique URLs loaded")

    existing = load_json(OUTPUT_FILE)
    existing_ids = {job["job_id"] for job in existing if job.get("job_id")}
    print(f"  {len(existing)} existing enriched jobs ({len(existing_ids)} unique IDs)")

    enriched = list(existing)
    new_count = 0
    scraped_ok = 0
    scraped_fail = 0
    total = len(excel_rows)
    skipped_existing = 0

    for i, row in enumerate(excel_rows, 1):
        if _interrupted:
            break

        jid = extract_job_id(row["url"])

        if jid and jid in existing_ids:
            skipped_existing += 1
            continue

        if i % 10 == 0 or i == 1:
            print(f"\n[{i}/{total}] {row['url']}")

        html = fetch_page(row["url"])
        if html:
            detail = parse_detail_page(html)
            ok = True
            scraped_ok += 1
            if i % 10 == 0:
                print(f"  ✓ {detail['job_title'][:70]}")
        else:
            detail = None
            ok = False
            scraped_fail += 1
            if i % 10 == 0:
                print(f"  ✗ unavailable (using Excel data)")

        job = build_enriched_job(row, detail, ok)
        enriched.append(job)
        if jid:
            existing_ids.add(jid)
        new_count += 1

        if new_count % SAVE_EVERY == 0:
            save_json(OUTPUT_FILE, enriched)
            print(f"  [saved {len(enriched)} jobs — {scraped_ok} scraped, {scraped_fail} failed, {skipped_existing} skipped]")

        if i < total:
            time.sleep(DELAY)

    save_json(OUTPUT_FILE, enriched)

    print(f"\n{'='*60}")
    print(f"Import complete.")
    print(f"  New jobs added:       {new_count}")
    print(f"  Successfully scraped: {scraped_ok}")
    print(f"  Failed (used Excel):  {scraped_fail}")
    print(f"  Already existed:      {skipped_existing}")
    print(f"  Total in {OUTPUT_FILE}: {len(enriched)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
